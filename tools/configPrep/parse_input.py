"""
Create configuration files for dbDIPview package metadata folder

Input:
	xls 
Output:
	createdb.sql
	queries.xml
	list.xml
"""

import openpyxl
import one_screen
import constants as cst

foreign_keys = []
all_schemas = []
all_tables_and_csv_files = {}


def cell_value(worksheet, row_num, col_num):
	cell = worksheet.cell(row=row_num, column=col_num)
	out = cell.value
	if out:
		return cell.value
	else:
		return ""


def xlsc2queriesxml(file, sheet_index, my_XML_output, my_createdb, my_createdb01, pass_num):
	# copy excel into csv file

	global foreign_keys, all_tables_and_csv_files
	schema = "schema_unknown"
	ctable = ""
	current_state = 0
	excel_row_number = 0
	usedSifranti = []
	a_joins = []

	xcb = one_screen.XML_query_block(pass_num)

	workbook = openpyxl.load_workbook(file)
	worksheet = workbook.worksheets[sheet_index]
	num_rows = worksheet.max_row
	current_state = cst.WAIT_ID

	#print("----sheet index:" + str(sheet_index) + " ---num of rows:" + str(num_rows))
	for curr_row in range(1, num_rows+1):
		excel_row_number += 1

		current_label = cell_value(worksheet, curr_row, 1)           #worksheet.cell_value(curr_row, 0)
		ccolumn = str(cell_value(worksheet, curr_row, 2))    #str(worksheet.cell_value(curr_row, index_col))
		#cell = worksheet.cell(curr_row, index_col)
		rowColumns = worksheet.max_column #row_len(curr_row)

		if (current_label or ccolumn) and current_state != cst.SKIP_ALL:
			#print("-------------------------line " + str(excel_row_number))
			#print(" label:" + str(current_label) + " value:" + ccolumn)
			#print(" NUM. OF COLUMNS:" + str(rowColumns))
			#print(" STATE:" + str(current_state))
			if not current_label and not ccolumn:
				print("  PASS")
				pass

			elif current_state == cst.WAIT_ID:
				check_label(sheet_index, excel_row_number, current_label, "Id:")
				current_id = ccolumn.strip()
				xcb.elements_id(current_id)
				current_state = current_state + 1
				print(" ID=" + ccolumn)

			elif current_state == cst.WAIT_SELECTDESC:
				check_label(sheet_index, excel_row_number, current_label, "SelectDescription:")
				xcb.elements_select_description(ccolumn.strip())
				current_state = current_state + 1
				#print(" SELECTDESC=" + str(ccolumn))

			elif current_state == cst.WAIT_TITLE:
				check_label(sheet_index, excel_row_number, current_label, "Title:")
				xcb.elements_set_title(ccolumn.strip())
				current_state = current_state + 1
				#print(" TITLE=" + str(ccolumn))

			elif current_state == cst.WAIT_SUBTITLE or current_state == cst.WAIT_SUBTITLE_SUBSELECT:
				check_label(sheet_index, excel_row_number, current_label, "Subtitle:")
				if current_state == cst.WAIT_SUBTITLE:
					xcb.elements_subtitle(ccolumn.strip())
				else:
					xcb.elements_subtitle_subselect(ccolumn.strip())
				current_state = current_state + 1
				#print(" SUBTITLE=" + str(ccolumn))

			elif current_state == cst.WAIT_SCHEMA or current_state == cst.WAIT_SCHEMA_SUBSELECT:
				check_label(sheet_index, excel_row_number, current_label, "Schema:")
				schema = str(ccolumn)
				xcb.elements_set_schema(schema.strip())
				current_state = current_state + 1
				#print(" SCHEMA=" + schema)
				if schema not in all_schemas:
					all_schemas.append(schema)

			elif current_state == cst.WAIT_TABLE_MAIN:
				check_label(sheet_index, excel_row_number, current_label, "Table:")
				ctable = ccolumn.strip()
				my_createdb.begin_table(schema, ctable)
				my_createdb01.begin_table(schema, ctable)
				current_state = current_state + 1
				#print(" TABLE_MAIN: table=" + schema + "." + ctable)

			elif current_state == cst.WAIT_CSV_FILE or current_state == cst.WAIT_CSV_FILE_SUBSELECT:
				check_label(sheet_index, excel_row_number, current_label, "CSV file:")
				csv_file = ccolumn.strip()
				current_state = current_state + 1
				#print(" CSV FILE:" + csv_file)
				if csv_file not in all_tables_and_csv_files:
					all_tables_and_csv_files[csv_file] = schema + "." + ctable

			elif current_state == cst.WAIT_TITLE_SUBSELECT:
				check_label(sheet_index, excel_row_number, current_label, "Title:")
				xcb.elements_set_title_subselect(ccolumn.strip())
				current_state = current_state + 1
				#print(" TITLE_SUBSELECT=" + str(ccolumn))

			elif current_state == cst.WAIT_TABLE_SUBSELECT:
				check_label(sheet_index, excel_row_number, current_label, "Table:")
				current_state = current_state + 1
				ctable = ccolumn.strip().replace(".csv", "") 
				my_createdb.begin_table(schema, ctable)
				notFirstQuery = True
				#print(" TABLE_SUBSELECT current table: " +  schema + "." + ctable)

			elif current_state == cst.WAIT_HEADER or current_state == cst.WAIT_HEADER_SUBSELECT:
				a_joins = []
				if current_state == cst.WAIT_HEADER:
					xcb.elements_query_start()
				else:
					xcb.elements_query_start_subselect()
				#print(" HEADER=" +  str(ccolumn) + " (table header is ignored)")
				current_state = current_state + 1

			elif current_state == cst.WAIT_NEXT_FIELD or current_state == cst.WAIT_NEXT_FIELD_SUBS:
				#table data after table header
				if rowColumns < 11:
					print("Error, not enough columns in input, only: " + str(rowColumns) + ", sheet #" + str(sheet_index+1))
					exit()
				IN_COL_AS = 3
				IN_COL_TYPE = 4
				IN_COL_PRIMARY_KEY = 5
				IN_COL_KEYVALUE = 6
				IN_COL_KEYVALUE_KEY = 7
				IN_COL_KEYVALUE_VALUE = 8
				IN_COL_LABEL = 9
				IN_COL_LINK_TO_NEXT_SCREEN = 10
				IN_COMMENT_ON_COLUMN = 11
				ccolumn_as =          cell_value(worksheet, curr_row, IN_COL_AS)
				ccolumn_type =        cell_value(worksheet, curr_row, IN_COL_TYPE)
				ccolumn_pk =          cell_value(worksheet, curr_row, IN_COL_PRIMARY_KEY)
				dict =                cell_value(worksheet, curr_row, IN_COL_KEYVALUE)
				dict_in_column =      cell_value(worksheet, curr_row, IN_COL_KEYVALUE_KEY)
				dict_out_column =     cell_value(worksheet, curr_row, IN_COL_KEYVALUE_VALUE)
				ccolumn_input_label = cell_value(worksheet, curr_row, IN_COL_LABEL)
				ccolumn_ltns_target = cell_value(worksheet, curr_row, IN_COL_LINK_TO_NEXT_SCREEN)
				comment_on_column =   cell_value(worksheet, curr_row, IN_COMMENT_ON_COLUMN)

				#print("   FIELD=" +  str(ccolumn) + " state=" + str(current_state))

				my_createdb.one_column(ccolumn, ccolumn_as, ccolumn_type, ccolumn_pk)
				my_createdb01.one_column(schema, ctable, ccolumn, comment_on_column)

				if dict_in_column != "0" and dict_in_column != "": 
					usedSifranti.append(dict + "." + dict_in_column)

				alter_table = xcb.element_field(ctable, ccolumn, ccolumn_as, dict, dict_in_column, dict_out_column, a_joins)
				if my_createdb.is_table_active():
					foreign_keys.append(alter_table)

				if current_state == cst.WAIT_NEXT_FIELD:
					xcb.elements_param_long(ccolumn_input_label, ctable, ccolumn, ccolumn_type, dict, dict_in_column, dict_out_column, ccolumn_ltns_target)
					# TODO: expects current column
					if ccolumn_ltns_target:
						if ccolumn_as:
							cc = ccolumn_as
						else:
							cc = ccolumn
						xcb.elements_set_link_to_next_screen(cc, current_id, ctable, ccolumn)

				if current_state == cst.WAIT_NEXT_FIELD_SUBS:
					#xcb.elements_param_in_pass1(ccolumn_input_label, ctable, ccolumn, ccolumn_type, dict, dict_in_column, dict_out_column, ccolumn_ltns_target)
					if ccolumn_ltns_target:
						xcb.elements_param_in_subselect(ctable, ccolumn)

			elif current_state == cst.SKIP_ALL:
				print("SKIP_ALL AFTER MAIN TABLE")

			else:
				print("UNKNOW STATE: " + str(current_state))

		else:
			#print("EMPTY LINE or SKIP_ALL, pass=" + str(pass_num) + " cstate=" + str(current_state))

			if  current_state == cst.WAIT_NEXT_FIELD or current_state == cst.WAIT_NEXT_FIELD_SUBS:
				end_of_one_table(False, ctable, a_joins, my_createdb, my_createdb01, xcb, pass_num, my_XML_output, current_state, sheet_index)
				usedSifranti.clear()

				if pass_num == 1:
					current_state = cst.SKIP_ALL

				if current_state == cst.WAIT_NEXT_FIELD:
					xcb.clear_select()
					current_state = current_state + 1

				if current_state == cst.WAIT_NEXT_FIELD_SUBS:
					xcb.clear_select()
					current_state = cst.WAIT_TITLE_SUBSELECT


	#print("TABLE ENDED, pass=" + str(pass_num) + " cstate=" + str(current_state))
	if current_state != cst.SKIP_ALL:
		end_of_one_table(True, ctable, a_joins, my_createdb, my_createdb01, xcb, pass_num, my_XML_output, current_state, sheet_index)
		
	workbook.close()


def end_of_one_table(eof, ctable, a_joins, my_createdb, my_createdb01, xcb, pass_num, my_XML_output, cstate, sheet):
	my_createdb.end_table()
	my_createdb01.end_table()
	if cstate == cst.WAIT_NEXT_FIELD or cstate == cst.WAIT_NEXT_FIELD_SUBS:
		xcb.select_from(ctable, a_joins)

	if pass_num == 1:
		xcb.elements_screenend(ctable, sheet)
	else:
		if cstate == cst.WAIT_NEXT_FIELD:
			xcb.elements_query_end_pass2()
			if eof:
				xcb.elements_screen_end_pass2()
		else:
			#xcb.debug("STATE="+str(cstate))
			if cstate != cst.WAIT_TITLE_SUBSELECT:
				xcb.elements_subselect_end()
			if eof:
				xcb.elements_screen_end_pass2()

	my_XML_output.write(xcb.flush())


def check_label(sheet, row, clabel, label):
	if clabel != label:
		print("Error, sheet #" + str(sheet+1) + " row " + str(row) + ": expecting label '" + label + "', but got: '" + clabel + "'")
		exit()
